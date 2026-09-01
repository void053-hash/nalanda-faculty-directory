import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_EXCEL_FILE, ENRICHED_FACULTY_FILE

logging.basicConfig(level=logging.INFO, format="%(asctime)s - [%(levelname)s] - %(message)s")
logger = logging.getLogger("ExcelExporter")


def flatten_faculty_record(fac: Dict[str, Any]) -> Dict[str, Any]:
    """Flattens a rich faculty dictionary into a tabular row for Excel export."""
    name = fac.get("name", "")
    desig = fac.get("designation", "")
    dept = fac.get("department", "")
    email = fac.get("email", "")
    citations = fac.get("total_citations", 0)
    works = fac.get("total_works", 0)
    topics = ", ".join(fac.get("core_topics", []))
    methods = fac.get("methodologies_used", "")
    focus = fac.get("research_focus", "")
    reach_out = fac.get("student_reach_out_summary", "")
    profile_url = fac.get("profile_url", "")
    
    top_papers = fac.get("top_papers", [])
    paper1 = f"[{top_papers[0].get('year', '')}] {top_papers[0].get('title', '')} (Citations: {top_papers[0].get('citations', 0)})" if len(top_papers) > 0 else "N/A"
    paper2 = f"[{top_papers[1].get('year', '')}] {top_papers[1].get('title', '')} (Citations: {top_papers[1].get('citations', 0)})" if len(top_papers) > 1 else "N/A"
    paper3 = f"[{top_papers[2].get('year', '')}] {top_papers[2].get('title', '')} (Citations: {top_papers[2].get('citations', 0)})" if len(top_papers) > 2 else "N/A"

    return {
        "Faculty Name": name,
        "Designation": desig,
        "School / Department": dept,
        "Email Contact": email,
        "Total Citations": citations,
        "Total Works": works,
        "Core Research Topics": topics,
        "Methodologies Used": methods,
        "Research Focus": focus,
        "Student Reach-Out Advice": reach_out,
        "Recent Paper 1": paper1,
        "Recent Paper 2": paper2,
        "Recent Paper 3": paper3,
        "Profile URL": profile_url
    }


def style_excel_workbook(file_path: Union[str, Path]) -> None:
    """Applies professional typography, header colors, auto-widths, and borders to the Excel workbook."""
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "Faculty Academic Directory"

    # Styling definitions
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Royal Nalanda Navy
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Style header row
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        ws.row_dimensions[row[0].row].height = 55  # comfortable reading height
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            # Numeric / short columns center aligned
            if cell.column in [5, 6]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Column width rules
    col_width_overrides = {
        "A": 22, # Faculty Name
        "B": 20, # Designation
        "C": 28, # School / Department
        "D": 25, # Email Contact
        "E": 14, # Citations
        "F": 14, # Works
        "G": 30, # Topics
        "H": 32, # Methodologies
        "I": 40, # Research Focus
        "J": 45, # Student Advice
        "K": 35, # Paper 1
        "L": 35, # Paper 2
        "M": 35, # Paper 3
        "N": 30  # Profile URL
    }

    for col_letter, width in col_width_overrides.items():
        if col_letter in [get_column_letter(i) for i in range(1, ws.max_column + 1)]:
            ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"
    wb.save(file_path)
    logger.info(f"Excel styling applied successfully to {file_path}")


def export_to_excel(faculty_list: List[Dict[str, Any]], output_path: Union[str, Path] = OUTPUT_EXCEL_FILE) -> pd.DataFrame:
    """
    Main function for Module 4.
    Converts enriched faculty list to DataFrame and writes a styled Excel workbook.
    """
    flat_data = [flatten_faculty_record(f) for f in faculty_list]
    df = pd.DataFrame(flat_data)
    
    # Save enriched JSON for caching
    with open(ENRICHED_FACULTY_FILE, "w", encoding="utf-8") as f:
        json.dump(faculty_list, f, indent=2, ensure_ascii=False)

    # Save to Excel
    output_path = Path(output_path)
    df.to_excel(output_path, index=False, engine="openpyxl")
    
    # Apply openpyxl styling
    style_excel_workbook(output_path)
    logger.info(f"Export completed: {len(df)} records written to {output_path}")
    return df


if __name__ == "__main__":
    sample_faculty = [
        {
            "name": "Prof. (Dr.) Abhay Kumar Singh",
            "designation": "Professor & Dean",
            "department": "School of Historical Studies",
            "email": "aksingh@nalandauniv.edu.in",
            "total_citations": 320,
            "total_works": 24,
            "core_topics": ["Archaeology", "Maritime Trade"],
            "methodologies_used": "Archival Epigraphy, Material Culture Analysis",
            "research_focus": "Investigates maritime trade networks and ancient epigraphy.",
            "student_reach_out_summary": "Review his maritime trade publications before connecting.",
            "top_papers": [
                {"year": 2023, "title": "Indian Ocean Maritime Routes", "citations": 45}
            ],
            "profile_url": "https://nalandauniv.edu.in"
        }
    ]
    df_res = export_to_excel(sample_faculty)
    print("Export test successful!")
    print(df_res.head())
